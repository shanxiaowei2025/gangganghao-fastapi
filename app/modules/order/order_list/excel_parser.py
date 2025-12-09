import re
import openpyxl
from typing import List, Dict, Optional
from io import BytesIO


class ExcelParser:
    """Excel 文件解析器"""
    
    # 列名映射（Excel 列标题 -> 数据库字段）
    COLUMN_MAPPING = {
        '项目名称': 'project_name',
        '构件': 'component',
        '部位': 'location',
        '构件名称': 'component_name',
        '编号': 'number',
        '级直\n别径': 'level_diameter',
        '级别直径': 'level_diameter',
        '钢 筋 简 图': 'rebar_sketch',
        '钢筋简图': 'rebar_sketch',
        '图形信息': 'graph_info',
        '边角结构': 'edge_structure',
        '下料\n(mm)': 'cutting_length_mm',
        '下料\r\n(mm)': 'cutting_length_mm',
        '下料(mm)': 'cutting_length_mm',
        '根 件\n数*数': 'quantity_pieces',
        '根 件\r\n数*数': 'quantity_pieces',
        '根数*件数': 'quantity_pieces',
        '总\n根 数': 'total_quantity',
        '总\r\n根 数': 'total_quantity',
        '总根数': 'total_quantity',
        '重量\n(kg)': 'weight_kg',
        '重量\r\n(kg)': 'weight_kg',
        '重量(kg)': 'weight_kg',
        '备注': 'remark',
        # 兼容不同的列名变体
        '编 号': 'number',
        '级 别 直 径': 'level_diameter',
        '钢筋 简 图': 'rebar_sketch',
        '边 角 结 构': 'edge_structure',
        '下 料(mm)': 'cutting_length_mm',
        '根数 件数': 'quantity_pieces',
        '根 数*件 数': 'quantity_pieces',
        '总根 数': 'total_quantity',
        '重 量(kg)': 'weight_kg',
    }
    
    @staticmethod
    def normalize_text(text: str) -> str:
        """规范化文本：处理换行符和括号"""
        if not text:
            return text
        
        # 统一换行符（Windows \r\n 和 Linux \n）
        text = text.replace('\r\n', '\n')
        
        # 统一括号（中文括号转英文括号）
        text = text.replace('（', '(').replace('）', ')')
        
        # 移除多余的空格和换行符（但保留必要的）
        text = re.sub(r'\s+', ' ', text).strip()
        
        return text
    
    @staticmethod
    def extract_column_name(header: str) -> str:
        """提取列名，处理换行符和空格"""
        if not header:
            return ""
        
        # 统一换行符
        normalized = header.replace('\r\n', '\n')
        
        # 特殊处理：检查是否是图形信息字段
        if ExcelParser.is_graph_info_header(normalized):
            return '图形信息'
        
        # 尝试直接匹配
        if normalized in ExcelParser.COLUMN_MAPPING:
            return normalized
        
        # 尝试移除空格后匹配
        no_space = normalized.replace(' ', '')
        for key in ExcelParser.COLUMN_MAPPING:
            key_no_space = key.replace(' ', '')
            if key_no_space == no_space:
                return key
        
        # 尝试模糊匹配（包含关键字）
        for key in ExcelParser.COLUMN_MAPPING:
            key_no_space = key.replace(' ', '').replace('\n', '')
            normalized_no_space = normalized.replace(' ', '').replace('\n', '')
            if key_no_space == normalized_no_space:
                return key
        
        return normalized
    
    @staticmethod
    def is_graph_info_header(text: str) -> bool:
        """检查是否是图形信息字段的表头（前四个字是否为'图形信息'）"""
        if not text:
            return False
        
        # 移除空格和换行符后检查
        normalized = text.replace(' ', '').replace('\n', '').replace('\r', '')
        
        # 检查前四个字是否为'图形信息'
        return normalized.startswith('图形信息')
    
    @staticmethod
    def parse_graph_info(text: str) -> Optional[str]:
        """解析图形信息字段值（保留完整内容）"""
        if not text:
            return None
        
        # 直接返回规范化后的文本
        return ExcelParser.normalize_text(text)
    
    @staticmethod
    def find_header_row(worksheet) -> int:
        """
        自动查找表头行（包含"项目名称"或其他关键字段的行）
        
        Returns:
            表头行号（1-indexed）
        """
        key_fields = ['项目名称', '编号', '级别直径', '级直别径', '钢筋简图', '边角结构', '下料', '根数', '总根', '重量']
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=15, values_only=False), start=1):
            row_values = [str(cell.value) if cell.value else "" for cell in row]
            
            # 统计该行包含的关键字段数
            matched_count = 0
            for cell_value in row_values:
                if not cell_value or cell_value == "None":
                    continue
                    
                # 规范化单元格值
                normalized = cell_value.replace('\r\n', '\n').replace(' ', '')
                
                # 检查是否包含关键字段
                for key_field in key_fields:
                    key_normalized = key_field.replace(' ', '').replace('\n', '')
                    if key_normalized in normalized or normalized in key_normalized:
                        matched_count += 1
                        break
            
            # 如果该行包含至少3个关键字段，认为是表头行
            if matched_count >= 3:
                return row_idx
        
        # 默认第一行
        return 1
    
    @staticmethod
    def extract_value_after_colon(text: str) -> Optional[str]:
        """
        从"字段名：值"格式的文本中提取冒号后的值
        
        Args:
            text: 包含冒号的文本
            
        Returns:
            冒号后的值，如果没有冒号则返回原文本
        """
        if not text:
            return None
        
        # 处理中文冒号和英文冒号
        if '：' in text:
            parts = text.split('：', 1)
            return ExcelParser.normalize_text(parts[1]) if len(parts) > 1 else text
        elif ':' in text:
            parts = text.split(':', 1)
            return ExcelParser.normalize_text(parts[1]) if len(parts) > 1 else text
        
        return ExcelParser.normalize_text(text)
    
    @staticmethod
    def parse_excel(file_content: bytes) -> List[Dict]:
        """
        解析 Excel 文件
        
        Args:
            file_content: Excel 文件的二进制内容
            
        Returns:
            订单数据字典列表
        """
        try:
            # 读取 Excel 文件
            workbook = openpyxl.load_workbook(BytesIO(file_content))
            worksheet = workbook.active
            
            # 自动查找表头行
            header_row_idx = ExcelParser.find_header_row(worksheet)
            
            # 获取表头
            headers = []
            for cell in worksheet[header_row_idx]:
                if cell.value:
                    headers.append(ExcelParser.extract_column_name(str(cell.value)))
                else:
                    headers.append(None)
            
            # 构建列索引映射
            column_index_map = {}
            for idx, header in enumerate(headers):
                if header and header in ExcelParser.COLUMN_MAPPING:
                    field_name = ExcelParser.COLUMN_MAPPING[header]
                    column_index_map[idx] = field_name
            
            # 提取元数据（项目名称、构件）
            metadata = {}
            
            # 从 A2 提取项目名称
            a2_value = worksheet['A2'].value
            if a2_value:
                metadata['project_name'] = ExcelParser.extract_value_after_colon(str(a2_value))
            
            # 从 G2 提取构件
            g2_value = worksheet['G2'].value
            if g2_value:
                metadata['component'] = ExcelParser.extract_value_after_colon(str(g2_value))
            
            # 解析数据行（从表头行之后开始）
            data_list = []
            current_location = None  # 用于存储当前的"部位"
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row_idx + 1, values_only=False), start=header_row_idx + 1):
                # 检查是否是"部位"行（单独为一行的数据）
                row_values = [cell.value for cell in row]
                
                # 如果该行只有一个非空值，可能是部位标题行
                non_empty_count = sum(1 for v in row_values if v is not None and str(v).strip())
                
                if non_empty_count == 1:
                    # 找到非空值作为部位
                    for cell_value in row_values:
                        if cell_value and str(cell_value).strip():
                            current_location = ExcelParser.normalize_text(str(cell_value))
                            break
                    continue
                
                # 解析数据行
                row_data = {}
                
                for col_idx, cell in enumerate(row):
                    if col_idx in column_index_map:
                        field_name = column_index_map[col_idx]
                        cell_value = cell.value
                        
                        if cell_value is not None:
                            cell_value = ExcelParser.normalize_text(str(cell_value))
                        
                        # 特殊处理图形信息字段
                        if field_name == 'graph_info':
                            cell_value = ExcelParser.parse_graph_info(cell_value)
                        
                        row_data[field_name] = cell_value
                
                # 添加元数据（项目名称、构件）
                if 'project_name' in metadata and 'project_name' not in row_data:
                    row_data['project_name'] = metadata['project_name']
                if 'component' in metadata and 'component' not in row_data:
                    row_data['component'] = metadata['component']
                
                # 添加部位信息
                if current_location and 'location' not in row_data:
                    row_data['location'] = current_location
                
                # 只有当行中有实际数据时才添加
                if row_data:
                    data_list.append(row_data)
            
            # 过滤掉只有元数据（project_name、component、location）而没有实际订单数据的行
            filtered_data = []
            for item in data_list:
                # 检查是否有除了 project_name、component、location 之外的非空字段
                has_real_data = False
                for key, value in item.items():
                    if key not in ['project_name', 'component', 'location'] and value is not None:
                        has_real_data = True
                        break
                
                if has_real_data:
                    filtered_data.append(item)
            
            return filtered_data
        
        except Exception as e:
            raise ValueError(f"Excel 文件解析失败: {str(e)}")
